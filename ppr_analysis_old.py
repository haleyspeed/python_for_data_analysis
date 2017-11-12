__author__ = 'Haley E. Speed, Ph.D'
__copyright__ = 'copyright 2012'

# paired_pulse.py
# Last updated 5-25-15

import openpyxl
import sys
import os.path

from openpyxl import load_workbook
from sys import platform as _platform

# Routine to get the filename and path from the user
file_check = False

while file_check == False:

    user_path = raw_input("Enter the file path for the excel file. Type 'stop' to end the program")
    filepath = str (user_path)
    print "\n", "\t", "You entered", "'",filepath,"'","\n"

    if filepath == "stop" or filepath == "Stop" or filepath == "STOP":
        print "\t", "Program Stopped."
        sys.exit()

    elif filepath == "help" or filepath == "Help" or filepath == "HELP":

        # Help routine if the user is running windows
        if _platform == "win32":
            filepath = "Enter the file path for the excel file. Type 'stop' to end the program"

    # Validates that the filepath they entered leads to a valid filename. If not, it prompts them again.
    else:
        validate = os.path.isfile(filepath)
        if validate == True:
            file_check = True
            print "File is good"
        else:
            print "\t", "Nope. That path didn't work. Try again.", "\n", \
                "\t","Remember to include the filename at the end of the path (i.e. /Dropbox/Shank3/Ephys.xlsx)"

            print "\n"

wb = load_workbook(filepath)
print "\t", "Excel file opened successfully.", "\n"

input_sheet = raw_input("Enter the name of the worksheet (bottom tab)? or 'stop' to exit.")
sheet = input_sheet

print "\t", "You entered", sheet

if sheet == "stop" or sheet == "Stop" or sheet == "STOP":
    print "\t", "Program Stopped."
    sys.exit()
else:
    ws = wb [sheet]
    print "\t", "Sheet opened successfully.", "\n"
    print "\t", "Commencing analysis...."

max_col = ws.get_highest_column()
max_row = ws.get_highest_row ()
col = 1
norm_value = 0


# Saving Data
new_sheet =  sheet + "Analyzed"
ws2 = wb.create_sheet(title = new_sheet)

# Starts Analysis for one column
for col in range (1, max_col + 1):

    # Need to determine whether this is a ISI column, pulse 1 column, or pulse 2 column
    col_sub = (col-1)
    col_check = col_sub % 3 == 0
    ISI_row = 1
    PPR_row = 2
    max_col = ws2.get_highest_column()

    if col_check is True:
        if max_col == 1:
            out_col = 1
        else:
            out_col = ws2.get_highest_column() + 1

        # Saves Experiment name and ISIs in the new sheet
        for ISI_row in range (1, max_row + 1):
            ISI_value = ws.cell (row = ISI_row, column = col).value

            # Saves experiment name to the first row of the column of analyzed data
            _ = ws2.cell(column=out_col, row = ISI_row, value = ISI_value)

            ISI_row += 1

        # Adds "PPR" as title of Column 2
        _ = ws2.cell (column = out_col + 1, row = 1, value = "PPR")

        # Takes value of Col 3 and divides by column 2 to generate PPR
        for PPR_row in range (2, max_row + 1):
            pulse1_col = col + 1
            pulse2_col = col + 2

            pulse1_value = ws.cell (row = PPR_row, column = pulse1_col).value
            pulse2_value = ws.cell (row = PPR_row, column = pulse2_col).value
            PPR = pulse2_value / pulse1_value

            _ = ws2.cell (column = out_col + 1, row = PPR_row, value = PPR)

            PPR_row += 1

    col_check = False
    col += 1

# No more columns, so you're done
print "Fin."

# Save the File
wb.save(filepath)


print "Your PPR data has been added to the workbook as", sheet,"Analyzed."


sys.exit()



