__author__ = 'Haley E. Speed, Ph.D.'
__copyright__ = "Copyright 2012"
#LTP_Analysis.py

# Last Updated 5-24-15

import openpyxl
import sys
from openpyxl import load_workbook
from sys import platform as _platform
import os.path


# Routine to get the filename and path from the user
file_check = False

while file_check == False:

    user_path = raw_input("Enter the file path for the excel file. Type 'stop' to end the program")
    filepath = str (user_path)
    print "\n", "\t", "You entered", "'",filepath,"'","\n"

    if filepath == "stop" or filepath == "Stop" or filepath == "STOP":
        print "\t", "Program Stopped."
        sys.exit()

    # Validates that the filepath they entered leads to a valid filename. If not, it prompts them again.
    else:
        validate = os.path.isfile(filepath)
        if validate == True:
            file_check = True
            print "File is good"
        else:
            print "\t", "Nope. That path didn't work. Try again.", "\n", \
                "\t","Remember to include the filename at the end of the path (i.e. /Dropbox/Shank3/Ephys.xlslx)"

            print "\n"

wb = load_workbook(filepath)
print "\t", "Excel file opened successfully.", "\n"

input_sheet = raw_input("Enter the name of the worksheet (bottom tab)? or 'stop' to exit.")
sheet = str(input_sheet)

print "\t", "You entered", sheet

if sheet == "stop" or sheet == "Stop" or sheet == "STOP":
    print "\t", "Program Stopped."
    sys.exit()
else:
    ws = wb [sheet]
    print "\t", "Sheet opened successfully.", "\n"
    print "\t", "Commencing analysis...."

max_col = ws.get_highest_column()
max_row = ws.get_highest_row ()
col = 1
norm_value = 0

# Saving Data
new_sheet = "Norm " + sheet
ws2 = wb.create_sheet(title = new_sheet)

# Starts Analysis for one column
for col in range (1, max_col):
    baseline_sum = 0
    avg = 0
    row = 2
    print "column iteration", col
    base_row = 2

    #Start Analysis of raws within current column
    for row in range (2, 31):
        #Retrieve the value from the current cell
        get_value = ws.cell(row = row, column = col).value
        value_type = type (get_value)

        # Adds to the cumulative sum of raw data until row 32 or an empty cell (error in xlsx sheet setup)
        if get_value != "None" and type (get_value) is float:
            baseline_sum += get_value
            row += 1

        # When the end of the baseline rows is reached, The average is calculated
        else:
            print "There are empty cells in your data! Fix it and restart."
            sys.exit ()

    # Take Average from baseline and apply it to every cell in the column
    avg = baseline_sum/29
    norm_row = 2

    for norm_row in range (1,max_row):
        norm_value = 0
        get_value = ws.cell(row = norm_row, column = col).value

        # If get_value is not a number, it can't be averaged. This makes sure it's a number
        if get_value != "None" and norm_row != max_row and type (get_value) is float and norm_row != 1:
            norm_value = (get_value/avg) * 100
            print norm_value
            _ = ws2.cell(column=col, row=norm_row, value=norm_value)
            norm_row += 1

        # Specifies that row one is to keep current cell value
        elif norm_row == 1:
            _ = ws2.cell(column=col, row=1, value=get_value)


    #End of the column, now onto the next!
    else:
        print "\t", "Your baseline average for col", col, "is", avg


    col += 1

# No more columns, so you're done
print "Fin."

wb.save(filepath)


print "Your normalized data has been added to the workbook as 'Norm", sheet,"'"


sys.exit ()
